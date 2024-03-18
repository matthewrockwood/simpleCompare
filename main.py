import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def compare_csv_files(file1, file2, output_file):
    wb_output = Workbook()
    ws_output = wb_output.active

    with open(file1, 'r') as csvfile1, open(file2, 'r') as csvfile2:
        csvreader1 = csv.reader(csvfile1)
        csvreader2 = csv.reader(csvfile2)

        row_index = 1
        for row1, row2 in zip(csvreader1, csvreader2):
            col_index = 1
            for cell1, cell2 in zip(row1, row2):
                if cell1 != cell2:
                    diff_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    ws_output.cell(row=row_index, column=col_index).fill = diff_fill
                    ws_output.cell(row=row_index, column=col_index).value = f'{cell1} <> {cell2}'
                else:
                    ws_output.cell(row=row_index, column=col_index).value = cell1
                col_index += 1
            row_index += 1

    wb_output.save(output_file)
    print("Comparison completed. Differences highlighted in", output_file)


# Example usage:
file1 = "data.csv"
file2 = "data2.csv"
output_file = "comparison.xlsx"
compare_csv_files(file1, file2, output_file)
